'''
Created on Jan 24, 2017
@author: aliu
https://automatetheboringstuff.com/chapter12/

'''

file_path = "C:\\Users\\aliu\Box Sync\\aliu Cloud Drive\\SFDC_Data\\"
file_name = "Sample_Openpyxl.xlsx"
data_file = file_path + file_name
print (data_file)


import openpyxl

''' read the worksheet '''
input_wb = openpyxl.load_workbook(data_file)
input_wb.get_sheet_names()

Yr2015_data = input_wb.get_sheet_by_name('Yr2015')
Yr2016_data = input_wb.get_sheet_by_name('Yr2016')

''' read the cell'''
Yr2015_data['B1'].value
Yr2015_data.cell(row=1, column=2)
Yr2015_data.cell(row=1, column=2).value

for i in range(1, Yr2015_data.max_row):
    print (i, Yr2015_data.cell(row=i, column=3).value)
    
a = Yr2015_data.rows
for RowOfCellObjects in Yr2015_data.rows:
    for cellObj in RowOfCellObjects:
        print (cellObj.coordinate, cellObj.value)
    print('----End of Row ----')


for ColumnsofSheet in Yr2015_data.columns:
    for obj in ColumnsofSheet:
        print(obj.coordinate, obj.value)    
    print('End of Column')

for row in Yr2015_data.iter_rows():
    a = tuple(c.value for c in row)
    print (a)
  
    
import pandas as pd

''' Read column header and column
    Declare a empty dataframe
    to add column, df[name] = data
'''    
Yr2016_df = pd.DataFrame()
for col in Yr2016_data.columns:
    name = col[0].value
    d = list(c.value for c in col[1:])
    print (name, ";", d)
    print ('------- end of column --------')
    Yr2016_df[name] = d
Yr2016_df


'''http://pandas.pydata.org/pandas-docs/stable/tutorials.html'''
Yr2015_df = pd.read_excel(data_file,sheetname='Yr2015', header=None, names=['Index','Date','Product','Qty'])

Yr2016_df = pd.read_excel(data_file,sheetname='Yr2016', header=0)

sum(Yr2015_df.Qty)
min(Yr2015_df.Qty)
Yr2015_df.Qty.sum()
Yr2015_df.Qty.describe()
