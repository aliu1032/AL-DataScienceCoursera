'''
Created on Apr 19, 2017
@author: aliu

Purpose: Prepare the Appeal Case Data and Summary Data into an Excel
'''

import pandas as pd
import numpy as np


###########################################################
#   Current Claim Case & Appeal Status from OrderDetail   #
###########################################################
file_path = "C:\\Users\\aliu\\Box Sync\\aliu Cloud Drive\\Analytics\\Payor Analytics\\ALL-Apr19\\"
file_name="OrderDetailForAppeals_data.csv"
input_data = pd.read_csv(file_path+file_name, encoding='utf-8-sig')
input_data.columns = [cell.replace(' ','') for cell in input_data.columns]


prep_file_path = "C:\\Users\\aliu\\Box Sync\\aliu Cloud Drive\\Analytics\\Payor Analytics\\Scripting\\"
file_name = "GHI_OLI_Claim_Data_Prep.xlsx"
prep_note = pd.read_excel(prep_file_path+file_name, sheet_name = 'OrderDetail_for_Appeal', skiprows=1, encoding='utf-8-sig')
rename_columns = dict(zip(prep_note.OrderLineForAppeal, prep_note.Synonyms))

''' rename the columns that will be used for excel dashboard '''
input_data.rename(columns = rename_columns, inplace=True)


######################
#   Data Selection   #
######################
select_column = prep_note[(prep_note['Select Columns'] == 1)]['Synonyms']
ClaimCaseStatus_column = prep_note[(prep_note['ClaimCaseStatus'] == 1)]['Synonyms']

'''select a subset of Payor to be analysis'''
#select_payor = ["Anthem","United","Aetna","BC","Humana","Ontario","ServIT","British Columbia","Payor for"]
select_payor = ["United"]

# extend to unknown payor some time
select_payor_id = []

# retrieve the Payor Id and put the ids into a list
for payor in select_payor:
    a = input_data[(input_data['Tier2PayorName'].str.contains(payor, case=False))]['Tier2PayorID'].unique()
    b = input_data[(input_data['Tier2PayorID'].isin(a))]['Tier2PayorName'].unique()
    # print (a,b)
    select_payor_id = np.concatenate([select_payor_id,a])

#select by payor id
work_data = input_data[input_data['Tier2PayorID'].isin(select_payor_id)][select_column].copy()

del input_data
#select by Test deliver date
#work_data = input_data1[(input_data1['TestDeliveredDate'] >= '2016-10-1') &
#                              (input_data1['TestDeliveredDate'] <= '2016-11-30')]\
#                              [select_column]
                              
############################
#     Data Preparation     #
############################    

work_data['AppealDenialReasonCode'] = work_data['AppealDenialReasonCode'].astype(str)
work_data['QuadaxInsurancePlanCode'] = work_data['QuadaxInsurancePlanCode'].astype(str)
              
'''Assume Charge Error'''
work_data[['BilledAmount','TotalPayment','TotalAdjustment','TotalOutstanding']].fillna(0.0,inplace=True)
work_data['TotalChargeError'] = round(work_data['BilledAmount'] - work_data[['TotalPayment','TotalAdjustment','TotalOutstanding']].sum(1),2)

ClaimCaseStatus = work_data[ClaimCaseStatus_column]

############################
#     Reshape Data         #
############################



###############################################################
#     Completed Appeal Summary from QDX Appeal Summary Text   #
###############################################################

file_name="appealSuccess.txt"
input_data = pd.read_csv(file_path+file_name, encoding='utf-8-sig', delimiter = '|')


'''select a subset of Payor to be analysis
#select_payor = ["Anthem","United","Aetna","BC","Humana","Ontario","ServIT","British Columbia","Payor for"]
select_payor = ["Humana"]
'''
# extend to unknown payor some time
select_payor_id = []

# retrieve the Payor Id and put the ids into a list
for payor in select_payor:
    a = input_data[(input_data['appealInsCOName'].str.contains(payor, case=False))]['appealInsCOAltId'].unique()
    b = input_data[(input_data['appealInsCOAltId'].isin(a))]['appealInsCOName'].unique()
    #print (a,b)
    select_payor_id = np.concatenate([select_payor_id,a])


#select by payor id
CompletedAppealSummary = input_data[input_data['appealInsCOAltId'].isin(select_payor_id)].copy()

Unique_OLI = CompletedAppealSummary['appealAccession']
###############################################
#     Write the columns into a excel file     #
###############################################

file_path = "C:\\Users\\aliu\\Box Sync\\aliu Cloud Drive\\Analytics\\Payor Analytics\\UHC\\"
output_file = 'AppealData.xlsx'

writer = pd.ExcelWriter(file_path+output_file, engine='openpyxl')
ClaimCaseStatus.to_excel(writer, sheet_name='ClaimCase_Appeal_Status', index=False, encoding='utf-8')
CompletedAppealSummary.to_excel(writer, sheet_name='Completed_Appeal', index=False, encoding='utf-8')
writer.save()
writer.close()

string_a = 'Select * from [StagingDB].[ODS].[StgCaseStatus] where OrderLineItemID in '
from openpyxl import load_workbook
book = load_workbook(filename = file_path+output_file)
ws1 = book.create_sheet(title = 'ForSQL_CaseHistory')
ws1['A1'] = string_a + str(list(Unique_OLI))
#ws1 = book.get_sheet_by_name(book.get_sheet_names()[2])
#ws1['A1'] = str(list(Unique_OLI))
book.save(file_path+output_file)

